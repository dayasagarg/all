import openpyxl

def getRowCount(file,sheetName):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetName)
    return sheet.max_row

def getColumnCount(file,sheetName):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetName)
    return sheet.max_column

def readData(file,sheetName,rowNo,columnNo):
    wb =openpyxl.load_workbook(file,sheetName)
    sheet = wb.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowNo,column=columnNo).value

def writeData(file,sheetName,rowNo,columnNo,data):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheetName)
    sheet.cell(row=rowNo,column=columnNo).value = data
    wb.save(file)






